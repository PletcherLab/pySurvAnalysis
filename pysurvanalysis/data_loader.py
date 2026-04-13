"""Load and prepare individual-level survival data from experiment files.

Supported input formats
-----------------------

**Excel workbook (.xlsx)** — census-level data across three sheets:

* **RawData** sheet — one row per chamber per census time, with columns:
    AgeH, Chamber, IntDeaths, Censored  (at minimum)

* **Design** sheet — one row per chamber, with columns:
    Chamber, SampleSize, StartTime, <factor1>, <factor2>, ...
    All columns after *StartTime* are treatment factors.

* **PrivateData** sheet (optional) — configuration:
    AssumeCensored  (1=True, 0=False; default True)

**CSV / TSV (.csv, .tsv)** — individual-level data in one of two shapes:

* **Long format** (one row per individual):
    time_col, event_col, factor1_col, factor2_col, ...
    Default column names: ``Age`` (time), ``Event`` (event).
    Remaining columns are treated as factor columns unless ``factor_cols``
    is specified explicitly.

* **Wide ragged format** (one column per group-×-event combination):
    Each column contains survival times for one (factor_level_combo, event)
    group.  Columns can be different lengths.  Column-to-group mapping is
    either inferred from column names or supplied via ``col_mapping``.

Output
------
All loaders return a tuple ``(individual_df, factor_names)`` where
``individual_df`` has one row per *individual* with columns:

    time        — age at death or censoring (hours / same unit as input)
    event       — 1 = observed death, 0 = right-censored
    chamber     — source chamber id ("N/A" for CSV inputs)
    treatment   — combined treatment label (factor levels joined with "/")
    <factor1>, <factor2>, ...  — individual factor columns
"""

from __future__ import annotations

from pathlib import Path
from typing import Union

import pandas as pd


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def read_assume_censored(path: Union[str, Path]) -> bool:
    """Read the AssumeCensored flag from the PrivateData sheet.

    Returns True if AssumeCensored == 1, False otherwise.
    Falls back to True if the sheet or column is missing.
    """
    try:
        priv = pd.read_excel(path, sheet_name="PrivateData")
        if "AssumeCensored" in priv.columns and len(priv) > 0:
            return int(priv["AssumeCensored"].iloc[0]) == 1
    except (ValueError, KeyError):
        pass
    return True


def load_chamber_flags(path: Union[str, Path]) -> set:
    """Return the set of excluded chamber IDs from the ChamberFlags sheet.

    Reads the ``Excluded`` column; any chamber with value 1 is excluded.
    Returns an empty set if the sheet or column is missing.
    """
    try:
        cf = pd.read_excel(path, sheet_name="ChamberFlags")
        if "Chamber" in cf.columns and "Excluded" in cf.columns:
            excluded = cf.loc[cf["Excluded"] == 1, "Chamber"]
            return set(excluded.tolist())
    except (ValueError, KeyError):
        pass
    return set()


def load_defined_plots(path: Union[str, Path]) -> list[tuple[str, list[str]]]:
    """Read the DefinedPlots sheet and return one (name, treatments) pair per plot.

    Layout: **one column per plot**.

    * Row 1 of the column supplies the **plot name** (used in UI dropdowns and
      report headings).
    * Data rows start at Excel row 6.  Each cell names one treatment to include;
      only the text *before the first comma* is used.  An empty cell ends the
      list for that column.

    Returns a list of ``(name, treatment_list)`` tuples.  Returns an empty list
    if the sheet is absent or has no usable columns.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        if "DefinedPlots" not in wb.sheetnames:
            return []
        ws = wb["DefinedPlots"]

        plots: list[tuple[str, list[str]]] = []
        for col_cells in ws.iter_cols():
            # Row 1: plot name
            name_val = col_cells[0].value  # row 1 = index 0
            plot_name = str(name_val).strip() if name_val is not None else ""

            # Rows 6+: treatment labels (index 5 onward)
            treatments: list[str] = []
            for cell in col_cells[5:]:      # index 5 = Excel row 6
                val = cell.value
                if val is None or str(val).strip() == "":
                    break
                label = str(val).split(",")[0].strip()
                if label:
                    treatments.append(label)

            if treatments:
                plots.append((plot_name, treatments))
        return plots
    except Exception:
        return []


def load_design(path: Union[str, Path]) -> tuple[pd.DataFrame, list[str]]:
    """Read the Design sheet and return (design_df, factor_names).

    Factor columns are everything after the *StartTime* column.
    """
    design = pd.read_excel(path, sheet_name="Design")
    cols = list(design.columns)
    start_idx = cols.index("StartTime")
    factors = cols[start_idx + 1 :]
    if not factors:
        raise ValueError("No treatment factor columns found after StartTime in Design sheet")
    return design, factors


def load_raw_data(path: Union[str, Path]) -> pd.DataFrame:
    """Read the RawData sheet."""
    raw = pd.read_excel(path, sheet_name="RawData")
    required = {"AgeH", "Chamber", "IntDeaths", "Censored"}
    missing = required - set(raw.columns)
    if missing:
        raise ValueError(f"RawData sheet missing columns: {missing}")
    return raw


def build_individual_data(
    raw: pd.DataFrame,
    design: pd.DataFrame,
    factors: list[str],
    assume_censored: bool = True,
    excluded_chambers: set | None = None,
) -> pd.DataFrame:
    """Expand census-level counts into one row per individual.

    For each chamber at each census time:
    * IntDeaths > 0 → that many death rows (event=1)
    * Censored > 0  → that many censored rows (event=0)

    Parameters
    ----------
    assume_censored : bool
        If True (default), the initial cohort size is taken from the Design
        sheet's SampleSize column.  Individuals not accounted for by observed
        deaths or explicit censoring are assumed right-censored at the last
        census time and added to the dataset.

        If False, the initial cohort size for each chamber is calculated as
        the sum of all IntDeaths and Censored observations for that chamber.
        No additional right-censored individuals are added.
    excluded_chambers : set or None
        Chamber IDs to skip entirely (sourced from the ChamberFlags sheet).
        Excluded chambers contribute no rows to the output.
    """
    _excluded = excluded_chambers or set()
    rows: list[dict] = []

    # Pre-compute chamber → design info lookup
    design_lookup = design.set_index("Chamber")

    for chamber_id, grp in raw.groupby("Chamber"):
        if chamber_id in _excluded:
            continue
        grp = grp.sort_values("AgeH")

        if chamber_id not in design_lookup.index:
            continue

        chamber_design = design_lookup.loc[chamber_id]
        factor_vals = {f: chamber_design[f] for f in factors}

        accounted = 0  # total deaths + censored seen so far

        for _, row in grp.iterrows():
            age = row["AgeH"]
            n_deaths = int(row["IntDeaths"])
            n_censored = int(row["Censored"])

            for _ in range(n_deaths):
                rec = {"time": age, "event": 1, "chamber": chamber_id}
                rec.update(factor_vals)
                rows.append(rec)

            for _ in range(n_censored):
                rec = {"time": age, "event": 0, "chamber": chamber_id}
                rec.update(factor_vals)
                rows.append(rec)

            accounted += n_deaths + n_censored

        if assume_censored:
            # Remaining individuals are right-censored at the last census time
            n0 = int(chamber_design["SampleSize"])
            last_time = grp["AgeH"].max()
            remaining = n0 - accounted
            for _ in range(remaining):
                rec = {"time": last_time, "event": 0, "chamber": chamber_id}
                rec.update(factor_vals)
                rows.append(rec)

    df = pd.DataFrame(rows)

    # Build treatment label from factor columns
    df["treatment"] = df[factors].astype(str).agg("/".join, axis=1)

    # Order columns nicely
    col_order = ["time", "event", "chamber", "treatment"] + factors
    df = df[col_order].sort_values(["treatment", "time"]).reset_index(drop=True)

    return df


# ---------------------------------------------------------------------------
# CSV helpers
# ---------------------------------------------------------------------------

def _sanitize_token(value: str) -> str:
    """Lowercase alphanumeric representation of a string (for fuzzy matching)."""
    return "".join(ch.lower() for ch in str(value) if ch.isalnum())


def _infer_wide_column_mapping(
    raw_df: pd.DataFrame,
    factor_names: list[str],
    factor_levels: dict[str, list],
) -> dict[str, tuple[str, str, int]]:
    """Attempt to infer wide-format column → (level_a, level_b, event) mapping.

    Parameters
    ----------
    raw_df : the raw wide-format DataFrame
    factor_names : list of two factor names, e.g. ["IRS1", "Foxo"]
    factor_levels : dict mapping factor name → list of its levels

    Returns
    -------
    dict mapping column name → (level_factor0, level_factor1, event_value)
    """
    if len(factor_names) != 2:
        raise ValueError("Wide format auto-inference requires exactly 2 factors.")

    fa_name, fb_name = factor_names
    level_a_tokens = {lvl: _sanitize_token(str(lvl)) for lvl in factor_levels[fa_name]}
    level_b_tokens = {lvl: _sanitize_token(str(lvl)) for lvl in factor_levels[fb_name]}
    event_tokens = {
        1: ["event", "death", "dead", "died"],
        0: ["censored", "censor", "cens", "alive"],
    }

    mapping: dict[str, tuple[str, str, int]] = {}
    used_keys: set = set()

    for col in raw_df.columns:
        token = _sanitize_token(col)
        a_match = [lvl for lvl, lvl_token in level_a_tokens.items() if lvl_token and lvl_token in token]
        b_match = [lvl for lvl, lvl_token in level_b_tokens.items() if lvl_token and lvl_token in token]

        event_match = None
        for event_value, aliases in event_tokens.items():
            if any(alias in token for alias in aliases):
                event_match = event_value
                break

        if len(a_match) == 1 and len(b_match) == 1 and event_match in {0, 1}:
            key = (a_match[0], b_match[0], event_match)
            if key in used_keys:
                raise ValueError(f"Ambiguous wide mapping: multiple columns map to {key}.")
            used_keys.add(key)
            mapping[col] = key

    expected = len(factor_levels[fa_name]) * len(factor_levels[fb_name]) * 2
    if len(mapping) != expected:
        raise ValueError(
            f"Could not infer complete wide mapping from column names. "
            f"Mapped {len(mapping)} columns, expected {expected}. "
            "Provide explicit col_mapping."
        )
    return mapping


def _individual_df_from_rows(
    rows: list[dict],
    factors: list[str],
) -> pd.DataFrame:
    """Build a standardised individual-level DataFrame from a list of row dicts."""
    df = pd.DataFrame(rows)
    df["time"] = pd.to_numeric(df["time"], errors="coerce")
    df["event"] = pd.to_numeric(df["event"], errors="coerce").astype("Int64")
    df = df.dropna(subset=["time", "event"]).copy()
    df["event"] = df["event"].astype(int)
    if not set(df["event"].unique()).issubset({0, 1}):
        raise ValueError("Event column must contain only 0 and 1 values.")
    df["treatment"] = df[factors].astype(str).agg("/".join, axis=1)
    col_order = ["time", "event", "chamber", "treatment"] + factors
    df = df[col_order].sort_values(["treatment", "time"]).reset_index(drop=True)
    return df


def load_csv_long(
    path: Union[str, Path],
    time_col: str = "Age",
    event_col: str = "Event",
    factor_cols: list[str] | None = None,
) -> tuple[pd.DataFrame, list[str]]:
    """Load individual-level CSV/TSV data in long format.

    Parameters
    ----------
    path : path to .csv or .tsv file
    time_col : name of the survival-time column (default ``"Age"``)
    event_col : name of the event-indicator column (default ``"Event"``)
    factor_cols : list of factor column names.  If ``None``, all columns that
        are not ``time_col`` or ``event_col`` are treated as factors.

    Returns
    -------
    (individual_df, factor_names)  — same shape as ``load_experiment()``
    """
    path = Path(path)
    sep = "\t" if path.suffix.lower() == ".tsv" else ","
    raw = pd.read_csv(path, sep=sep)

    missing = [c for c in [time_col, event_col] if c not in raw.columns]
    if missing:
        raise ValueError(f"Long-format CSV missing columns: {missing}")

    if factor_cols is None:
        factor_cols = [c for c in raw.columns if c not in {time_col, event_col}]
    if not factor_cols:
        raise ValueError("No factor columns found in CSV. Specify factor_cols explicitly.")

    missing_factors = [f for f in factor_cols if f not in raw.columns]
    if missing_factors:
        raise ValueError(f"Factor columns not found in CSV: {missing_factors}")

    rows = []
    for _, row in raw.iterrows():
        rec: dict = {
            "time": row[time_col],
            "event": row[event_col],
            "chamber": "N/A",
        }
        for f in factor_cols:
            rec[f] = row[f]
        rows.append(rec)

    df = _individual_df_from_rows(rows, factor_cols)
    return df, factor_cols


def load_csv_wide(
    path: Union[str, Path],
    factor_names: list[str],
    factor_levels: dict[str, list] | None = None,
    col_mapping: list[dict] | None = None,
) -> tuple[pd.DataFrame, list[str]]:
    """Load individual-level data from a wide ragged CSV/TSV.

    Parameters
    ----------
    path : path to .csv or .tsv file
    factor_names : list of two factor names, e.g. ``["Sex", "Density"]``
    factor_levels : dict mapping factor name → list of its levels.
        Required when ``col_mapping`` is None (for auto-inference).
    col_mapping : explicit list of dicts, each with keys:
        ``column`` (str), ``factor1_level``, ``factor2_level``, ``event`` (0 or 1).
        If None, auto-inferred from column names.

    Returns
    -------
    (individual_df, factor_names)
    """
    if len(factor_names) != 2:
        raise ValueError("Wide format requires exactly 2 factor names.")

    path = Path(path)
    sep = "\t" if path.suffix.lower() == ".tsv" else ","
    raw = pd.read_csv(path, sep=sep)

    fa_name, fb_name = factor_names

    if col_mapping is not None:
        mapping: dict[str, tuple] = {}
        if not isinstance(col_mapping, list):
            raise ValueError("col_mapping must be a list of dicts.")
        for item in col_mapping:
            col = item.get("column")
            lvl_a = item.get("factor1_level")
            lvl_b = item.get("factor2_level")
            event = item.get("event")
            if col not in raw.columns:
                raise ValueError(f"Wide-format column not found: {col}")
            if event not in (0, 1):
                raise ValueError(f"Wide column '{col}' must set event to 0 or 1.")
            mapping[col] = (lvl_a, lvl_b, int(event))
    else:
        if factor_levels is None:
            raise ValueError("factor_levels is required when col_mapping is not provided.")
        mapping = _infer_wide_column_mapping(raw, factor_names, factor_levels)

    rows = []
    for col, (lvl_a, lvl_b, event) in mapping.items():
        for value in raw[col].dropna():
            rows.append({
                "time": float(value),
                "event": int(event),
                "chamber": "N/A",
                fa_name: lvl_a,
                fb_name: lvl_b,
            })

    if not rows:
        raise ValueError("No non-null observations found in wide-format columns.")

    df = _individual_df_from_rows(rows, factor_names)
    return df, factor_names


def detect_csv_format(
    path: Union[str, Path],
    time_col: str = "Age",
    event_col: str = "Event",
) -> str:
    """Heuristically detect whether a CSV is in long or wide format.

    Returns ``"long"`` or ``"wide"``.
    """
    path = Path(path)
    sep = "\t" if path.suffix.lower() == ".tsv" else ","
    raw = pd.read_csv(path, sep=sep, nrows=5)
    if time_col in raw.columns and event_col in raw.columns:
        return "long"
    return "wide"


# ---------------------------------------------------------------------------
# Unified high-level loader
# ---------------------------------------------------------------------------

def load_experiment(
    path: Union[str, Path],
    assume_censored: bool = True,
    excluded_chambers: set | None = None,
    # CSV-specific parameters
    time_col: str = "Age",
    event_col: str = "Event",
    factor_cols: list[str] | None = None,
    csv_format: str = "auto",
    col_mapping: list[dict] | None = None,
    factor_names: list[str] | None = None,
    factor_levels: dict[str, list] | None = None,
) -> tuple[pd.DataFrame, list[str]]:
    """High-level loader: returns (individual_data, factor_names).

    Routes to the correct loader based on file extension.

    Parameters
    ----------
    path : input file path (.xlsx, .csv, or .tsv)
    assume_censored : bool
        Excel only.  If True, unaccounted individuals (SampleSize minus
        observed deaths and censored) are added as right-censored at the
        last census time.  No effect for CSV inputs (already individual-level).
    excluded_chambers : set or None
        Excel only.  Chamber IDs to exclude entirely (from ChamberFlags).
        Caller is responsible for loading these via ``load_chamber_flags()``.
    time_col : CSV only.  Column name for survival time (default ``"Age"``).
    event_col : CSV only.  Column name for event indicator (default ``"Event"``).
    factor_cols : CSV long format.  Factor column names; auto-detected if None.
    csv_format : ``"auto"`` (default), ``"long"``, or ``"wide"``.
    col_mapping : CSV wide format.  Explicit column→group mapping list.
    factor_names : CSV wide format.  Factor names (required if wide + no mapping).
    factor_levels : CSV wide format.  Factor name → list of levels (for auto-inference).
    """
    path = Path(path)
    ext = path.suffix.lower()

    if ext == ".xlsx":
        design, factors = load_design(path)
        raw = load_raw_data(path)
        individual = build_individual_data(
            raw, design, factors,
            assume_censored=assume_censored,
            excluded_chambers=excluded_chambers,
        )
        return individual, factors

    if ext in {".csv", ".tsv"}:
        fmt = csv_format.lower() if csv_format else "auto"
        if fmt not in {"auto", "long", "wide"}:
            raise ValueError("csv_format must be 'auto', 'long', or 'wide'.")

        if fmt == "auto":
            fmt = detect_csv_format(path, time_col=time_col, event_col=event_col)

        if fmt == "long":
            return load_csv_long(path, time_col=time_col, event_col=event_col, factor_cols=factor_cols)

        # wide
        if factor_names is None:
            raise ValueError("factor_names is required for wide-format CSV loading.")
        return load_csv_wide(
            path,
            factor_names=factor_names,
            factor_levels=factor_levels,
            col_mapping=col_mapping,
        )

    raise ValueError(f"Unsupported file type: {ext!r}. Expected .xlsx, .csv, or .tsv.")
